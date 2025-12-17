# coding: UTF-8
from openai import OpenAI;
import openpyxl
import tiktoken
import json
import os, json, re
import numpy as np
import faiss
import spacy
_NLP = spacy.load("en_core_web_sm")

JSON_PATH = "need_oriented_environmental_knowledge_base.json" 

OPENAI_API_KEY = "OpenAI key" # ←Your API Key

EXCEL_NAME = "Evaluation_Needs_Rag.xlsx"
WORKBOOK_LIST = ["test_rag_border01","test_rag_border02","test_rag_border03","test_rag_border01_4o","test_rag_border02_4o","test_rag_border03_4o"]



#Token Count
All_token = 0

chat = []
script = []

#Read Knowledge Base
with open('need_oriented_environmental_knowledge_base.json') as f:
    needs_info = json.load(f)

prob_num=0
prob_list = []

situation = {'bathroom':['soap', 'bathroom cabinet', 'bathroom counter', 'bathtub','ceiling', 'ceiling lamp', 'closet', 'closet drawers', 'clothes pants', 'pile of clothes', 'clothes shirt', 'curtain', 'deodorant', 'door', 'door frame','face cream', 'faucet', 'floor', 'hair product', 'hanger', 'light switch','painkiller', 'perfume', 'power socket', 'rug', 'sink', 'slipper', 'private room','toilet', 'toilet paper', 'toothbrush', 'toothpaste', 'towel','towel rack', 'wall', 'wall lamp', 'wall picture frame', 'washing machine','cleaning sponge', 'Water glass', 'mirror'],
            'bedroom':['bed', 'books', 'cabinet', 'ceiling', 'ceiling lamp', 'chair','closet', 'closet drawers', 'clothes pants', 'pile of clothes','clothes shirt', 'deodorant', 'door', 'door frame', 'face cream','floor', 'hanger', 'light switch', 'nightstand', 'orchid', 'perfume','photo frame', 'pillow', 'power socket', 'radio', 'rug', 'sofa','table lamp', 'wall', 'wall lamp', 'wall picture frame', 'wall shelf'],
            'kitchen':['banana', 'bookshelf', 'box', 'slices of bread', 'ceiling', 'ceiling lamp','mobile phone', 'cereal', 'chair', 'chocolate syrup', 'pile of clothes','coffee maker', 'coffee pot', 'condiment bottle', 'condiment pendulum','cooking pot', 'cupcake', 'cutlet', 'deodorant', 'bowl of dishes','dishwasher', 'dishwashing liquid', 'door', 'door frame', 'face cream', 'faucet','floor', 'folder', 'refrigerator', 'kitchen cabinet', 'kitchen counter','kitchen table', 'light switch', 'microwave', 'minced meat', 'mug','orchid', 'oven tray', 'perfume', 'photo frame', 'pie', 'pillow', 'dish','power socket', 'pudding', 'rug', 'sink', 'sofa', 'stove', 'ventilation fan','toaster', 'TV', 'TV stand', 'wall', 'wall phone', 'picture frame on the wall','wall shelf', 'cleaning sponge', 'wine glass','table'],
            'living room':['book', 'ceiling', 'ceiling lamp', 'mobile phone', 'chair', 'closet','closet drawer', 'clothes pants', 'pile of clothes', 'clothes shirt','computer','cpu screen', 'cupcake', 'deodorant', 'desk', 'door frame','face cream', 'floor', 'hair product', 'hanger', 'keyboard', 'light switch','mouse', 'mouse pad', 'mug', 'nightstand', 'orchid','photo frame', 'pillow', 'dish', 'power socket', 'remote control', 'rug', 'sofa','table lamp', 'television', 'television stand', 'wall', 'wall lamp','wall picture frame', 'wall shelf', 'water glass', 'wine glass', 'garbage', 'scissors', 'electric fan', 'tissue']
             }

def calc_token(pronpt):
    encoding = tiktoken.get_encoding("o200k_base")
    num_tokens = len(encoding.encode(pronpt))
    return num_tokens

# Reading Excel files
def read_xl(wb,prob_list,needs_dict,WORKBOOK_NAME):

    ws = wb[WORKBOOK_NAME]

    for i in range(ws.max_row-1):
        prob_list.append(ws.cell(row=2+i,column=1).value)

    for j in needs_info:
        needs_dict[j["name"]] = j["things"].split(",")
        
# Writing to Excel files + Calculating evaluation values
def write_xl(wb,needs,response_things,token_all,WORKBOOK_NAME,num):

    ws = wb[WORKBOOK_NAME]

    ws.cell(row=2+num,column=4).value = needs    
    ws.cell(row=2+num,column=5).value = response_things

    ans_needs = ws.cell(row=2+num,column=2).value.split(",")
    ans_thing = ws.cell(row=2+num,column=3).value.split(",")
    needs_list = needs.split(",")
    thing_list = response_things.split(",")

    count = 0
    for needss in needs_list:
        if needss in ans_needs:
            count+=1
    ws.cell(row=2+num,column=7).value = round(count/len(needs_list),3)
    count = 0
    for ans_d in ans_needs:
        if ans_d in needs_list:
            count+=1
    ws.cell(row=2+num,column=8).value = round(count/len(ans_needs),3)    
    
    count = 0
    for things in thing_list:
        if things in ans_thing:
            count+=1
    ws.cell(row=2+num,column=10).value = round(count/len(thing_list),3)
    count = 0
    for ans_t in ans_thing:
        if ans_t in thing_list:
            count+=1
    ws.cell(row=2+num,column=11).value = round(count/len(ans_thing),3)

    ws.cell(row=2+num,column=12).value = token_all

    room_navi=[]
    for t in thing_list:
        if t in situation['bathroom']:
            room_navi.append("bathroom")
        if t in situation['bedroom']:
            room_navi.append("bedroom")
        if t in situation['kitchen']:
            room_navi.append("kitchen")    
        if t in situation['living room']:
            room_navi.append("living room")
    room_navi = set(room_navi)

    room_name = ""
    for r in room_navi:
        if room_name != "":
            room_name = room_name +','
        room_name = room_name + r

    ws.cell(row=2+num,column=6).value = room_name

    wb.save(EXCEL_NAME) 

#Extract the object and verb from the query
def _noun_phrase_for(token):

    if token.pos_ == "PRON":
        return ""  
    
    for chunk in token.doc.noun_chunks:
        if token.i >= chunk.start and token.i < chunk.end:
            return " ".join(t.lemma_ for t in chunk if not t.is_space)
        
    left_mods = [t for t in token.lefts if t.dep_ in ("compound", "amod", "nummod", "poss")]
    phrase = " ".join(t.lemma_ for t in left_mods + [token] if not t.is_space)
    return phrase

def _collect_objects(verb):
    """
    動詞 token から目的語候補（直接目的語・補語・前置詞目的語）を集める
    """
    objs = []

    # Direct object/complement
    for c in verb.children:
        if c.dep_ in ("dobj", "obj", "attr", "oprd") and c.pos_ in ("NOUN", "PROPN", "PRON"):
            objs.append(c)

    # Prepositional phrase → pobj
    for prep in (c for c in verb.children if c.dep_ == "prep"):
        for pobj in (c for c in prep.children if c.dep_ == "pobj" and c.pos_ in ("NOUN", "PROPN", "PRON")):
            objs.append(pobj)

    # Parallelism of objects (conj)
    extra = []
    for o in objs:
        extra.extend([c for c in o.conjuncts if c.pos_ in ("NOUN", "PROPN")])
    objs.extend(extra)

    return objs

def _phrasal_verb_lemma(verb):
    """
    句動詞（動詞 + particle）を 'pick up' のように結合して返す
    """
    parts = [verb.lemma_]
    prt = [c.text.lower() for c in verb.children if c.dep_ == "prt"]
    if prt:
        parts.extend(prt)
    return " ".join(parts)

def extract_verb_object_queries(text: str):
    """
    英文から 'verb object' 形式の短いフレーズ群を '; ' 連結で返す。
    例: "find a cup in kitchen and pick up bottle" -> "find cup; pick up bottle"
    """
    doc = _NLP(text)
    phrases = []

    for token in doc:
        if token.pos_ != "VERB":
            continue
        verb_str = _phrasal_verb_lemma(token)
        objs = _collect_objects(token)
        for obj in objs:
            obj_phrase = _noun_phrase_for(obj)
            if verb_str and obj_phrase:
                phrases.append(f"{verb_str} {obj_phrase}")

    seen = set()
    uniq = []
    for p in phrases:
        if p not in seen:
            uniq.append(p)
            seen.add(p)
    return "; ".join(uniq)


# Data Loading & Preprocessing
def load_needss(json_path=JSON_PATH):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    records = []
    for i, d in enumerate(data):
        name = d.get("name","")
        action = d.get("action","")
        explain = d.get("explain","")
        things = d.get("things","")
        text = f"need: {name}\naction: {action}\nthings: {things}\nabout: {explain}"
        records.append({
            "id": i,
            "name": name,
            "action": action,
            "explain": explain,
            "things": [t.strip() for t in things.split(",") if t.strip()],
            "text": text
        })
    return records

records = load_needss(JSON_PATH)


# Embedded function
def get_embeddings_openai(texts):
    # OpenAI: text-embedding-3-large
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY)
    resp = client.embeddings.create(
        model="text-embedding-3-large",
        input=texts
    )
    return np.array([d.embedding for d in resp.data], dtype="float32")


# Corpus Embedding & FAISS Indexing
corpus_texts = [r["text"] for r in records]
X = get_embeddings_openai(corpus_texts)

def l2_normalize(v):
    v = v / (np.linalg.norm(v, axis=1, keepdims=True) + 1e-12)
    return v
Xn = l2_normalize(X)

index = faiss.IndexFlatIP(Xn.shape[1])  
index.add(Xn)


# Similar Search Function
def embed_query(q):
        return l2_normalize(get_embeddings_openai([q]))[0]

def search_needs(query, top_k=20):

    # Create a query that extracts only the verb and object
    q_vo = extract_verb_object_queries(query.strip())
    print("クエリ" + q_vo)

    q_for_embed = (q_vo if q_vo else query).lower()
    qv = embed_query(q_for_embed).reshape(1, -1)
    D, I = index.search(qv, top_k)
    hits = []
    for score, idx in zip(D[0], I[0]):
        if idx == -1: 
            continue
        rec = records[idx]
        hits.append({
            "score": float(score),
            "name": rec["name"],
            "explain": rec["explain"],
            "action": rec["action"],
            "things": rec["things"]
        })
    return hits

def create_thing_list(needs_list,needs_dict):
    needs_thing = []
    things = ""
    for i in needs_dict.keys():
        if i in needs_list:
            needs_thing = needs_thing + needs_dict[i]
    needs_thing = set(needs_thing)
    for d in needs_thing:
        things += d + ","
    things = things[:-1]
    print("物リスト"+things)
    return things

def chat_response(problem,needs_thing):

    # Create a list of needs from requirements
    chat = []

    client = OpenAI(api_key=OPENAI_API_KEY)
   
    content = problem + """
    \nBelow is a list of possible objects for the above request. Please narrow down and select only those that satisfy the user's request.
        Please output your response in the following format.
        Please do not include spaces, use only nouns, and answer in the notation provided.
        For nouns consisting of two words, insert a space.
        Also do not add [] or '.
        Example is in the below.
        Example:bed,sofa\n
        """+str(needs_thing)
    # print(content3)
    chat=[]
    chat.append({"role": "user", "content": content})
    # print(content3)
    print("トークン数(user)："+str(calc_token(content)))
    print("<System>")
    # print(chat)
    response = client.chat.completions.create(
        model="gpt-4o-mini",messages=chat
        # model="gpt-4o",messages=chat        
    )
    response_thing = response.choices[0].message.content.lstrip()
    print(response_thing)
    token_all = +calc_token(content)+calc_token(response_thing)
    print(token_all)
    return response_thing,token_all
    
def main():
    border_list = [0.1, 0.2, 0.3, 0.1, 0.2, 0.3]
    #Initialize
    wb = openpyxl.load_workbook(EXCEL_NAME) # Read Excel
    for W,b in zip(WORKBOOK_LIST,border_list):
        a=0
        needs_dict = {}
        prob_list = []
        read_xl(wb,prob_list,needs_dict,W)
        for problem in prob_list:
            print("<User>\n"+problem)
            needs_list = []
            needss = ""
            for h in search_needs(problem, top_k=20):
                if h["score"] > b:
                    needs_list.append(h["name"])
                    print(h["score"])
            if needs_list == []:
                needs_list.append("[none]")
                print("[none]")
            for d in needs_list:
                needss += d + ","
            needss = needss[:-1]
            things = create_thing_list(needs_list,needs_dict)
            response_thing,token_all = chat_response(problem,things)
            write_xl(wb,needss,response_thing,token_all,W,a)
            a+=1

if __name__ == "__main__":
    main()
