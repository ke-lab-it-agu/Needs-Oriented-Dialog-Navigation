# coding: UTF-8
from openai import OpenAI
import openpyxl
import tiktoken
import json

JSON_PATH = "need_oriented_environmental_knowledge_base.json"
OPENAI_API_KEY = "OpenAI key"  # ←Your API Key
EXCEL_NAME = "Evaluation_Needs_LLM.xlsx"

client = OpenAI(api_key=OPENAI_API_KEY)

# count token
All_token = 0

# prompt
chat = []
script = []

# read json
with open(JSON_PATH, encoding="utf-8") as f:
    needs_info = json.load(f)

prob_num = 0
prob_list = []

scene_num = 4

# Murray's psychogenic needs only
needs_all = [n["name"] for n in needs_info]
content_murray = f"""Let's assume you are the guide for this house.
First, the user request flows.
When it flows and you say "select the needs", infer a need that the user wants to fill.
Then you send a list of items that satisfy the need, and the user should choose the one that best meets the need.
The needs should be output from Murray's psychogenic needs.
Murray's psychogenic needs should be output with the following spelling.
When responding, please output only nouns as in the example.
Do not output anything other than the needs included in the list below
Murray's Psychogenic Needs:{", ".join(needs_all)}
Also, please output your answers in a format similar to the example.
Also do not add [] or '.
example:Thirst,Hunger
"""

# action name
needs_action = [n["name"] + ":" + n["action"] + "\n" for n in needs_info]
content_action = f"""Let's assume you are the guide for the house.
First, the user's request flows.
When it flows and I say "select the needs", infer a need that the user wants to fill.
Then you send a list of items that satisfy the need, and the user should choose the one that best meets the need.
The list is in the format “needs”:"keywords", where keywords are descriptions of needs.
Do not output anything other than the needs included in the list below
Murray's Psychogenic Needs:{", ".join(needs_action)}
Also, please output your answers in a format similar to the example.
Also do not add [] or '.
example:Thirst,Hunger
"""

# needs and its explanation
needs_explain = [n["name"] + ":" + n["explain"] + "\n" for n in needs_info]
content_explain = f"""Let's assume you are the guide for the house.
First, the user's request flows.
When it flows and you say "select the needs", infer a need that the user wants to fill.
Then you send a list of items that satisfy the need, and the user should choose the one that best meets the need.
The list is in the format [“needs”:"description"], where keywords are descriptions of needs.
Do not output anything other than the needs included in the list below.
Needs list:
{", ".join(needs_explain)}
Also, please output your answers in a format similar to the example.
Also do not add [] or '.
example:Thirst,Hunger
"""

situation = {
    'bathroom': ['soap', 'bathroom cabinet', 'bathroom counter', 'bathtub', 'ceiling', 'ceiling lamp', 'closet',
                 'closet drawers', 'clothes pants', 'pile of clothes', 'clothes shirt', 'curtain', 'deodorant', 'door',
                 'door frame', 'face cream', 'faucet', 'floor', 'hair product', 'hanger', 'light switch', 'painkiller',
                 'perfume', 'power socket', 'rug', 'sink', 'slipper', 'private room', 'toilet', 'toilet paper',
                 'toothbrush', 'toothpaste', 'towel', 'towel rack', 'wall', 'wall lamp', 'wall picture frame',
                 'washing machine', 'cleaning sponge', 'Water glass', 'mirror'],
    'bedroom': ['bed', 'books', 'cabinet', 'ceiling', 'ceiling lamp', 'chair', 'closet', 'closet drawers',
                'clothes pants', 'pile of clothes', 'clothes shirt', 'deodorant', 'door', 'door frame', 'face cream',
                'floor', 'hanger', 'light switch', 'nightstand', 'orchid', 'perfume', 'photo frame', 'pillow',
                'power socket', 'radio', 'rug', 'sofa', 'table lamp', 'wall', 'wall lamp', 'wall picture frame',
                'wall shelf'],
    'kitchen': ['banana', 'bookshelf', 'box', 'slices of bread', 'ceiling', 'ceiling lamp', 'mobile phone', 'cereal',
                'chair', 'chocolate syrup', 'pile of clothes', 'coffee maker', 'coffee pot', 'condiment bottle',
                'condiment pendulum', 'cooking pot', 'cupcake', 'cutlet', 'deodorant', 'bowl of dishes', 'dishwasher',
                'dishwashing liquid', 'door', 'door frame', 'face cream', 'faucet', 'floor', 'folder', 'refrigerator',
                'kitchen cabinet', 'kitchen counter', 'kitchen table', 'light switch', 'microwave', 'minced meat',
                'mug', 'orchid', 'oven tray', 'perfume', 'photo frame', 'pie', 'pillow', 'dish', 'power socket',
                'pudding', 'rug', 'sink', 'sofa', 'stove', 'ventilation fan', 'toaster', 'TV', 'TV stand', 'wall',
                'wall phone', 'picture frame on the wall', 'wall shelf', 'cleaning sponge', 'wine glass', 'table'],
    'living room': ['book', 'ceiling', 'ceiling lamp', 'mobile phone', 'chair', 'closet', 'closet drawer',
                    'clothes pants', 'pile of clothes', 'clothes shirt', 'computer', 'cpu screen', 'cupcake',
                    'deodorant', 'desk', 'door frame', 'face cream', 'floor', 'hair product', 'hanger', 'keyboard',
                    'light switch', 'mouse', 'mouse pad', 'mug', 'nightstand', 'orchid', 'photo frame', 'pillow',
                    'dish', 'power socket', 'remote control', 'rug', 'sofa', 'table lamp', 'television',
                    'television stand', 'wall', 'wall lamp', 'wall picture frame', 'wall shelf', 'water glass',
                    'wine glass', 'garbage', 'scissors', 'electric fan', 'tissue']
}


def calc_token(prompt: str) -> int:
    encoding = tiktoken.get_encoding("o200k_base")
    return len(encoding.encode(prompt))


def read_xl(wb, prob_list, Need_dict, sheet):
    ws = wb[sheet]
    for i in range(ws.max_row - 1):
        prob_list.append(ws.cell(row=2 + i, column=1).value)

    # needs[things] dictionary -> get from json
    for j in needs_info:
        Need_dict[j["name"]] = j["things"].split(",")


def write_xl(wb, needs, thing, token_all, sheet, num):
    ws = wb[sheet]

    # write output
    ws.cell(row=2 + num, column=4).value = needs
    ws.cell(row=2 + num, column=5).value = thing

    # precision, recall
    ans_needs = ws.cell(row=2 + num, column=2).value.split(",")
    ans_thing = ws.cell(row=2 + num, column=3).value.split(",")
    needs_list = needs.split(",")
    thing_list = thing.split(",")

    count = 0
    for needss in needs_list:
        if needss in ans_needs:
            count += 1
    ws.cell(row=2 + num, column=7).value = round(count / len(needs_list), 3)

    count = 0
    for ans_d in ans_needs:
        if ans_d in needs_list:
            count += 1
    ws.cell(row=2 + num, column=8).value = round(count / len(ans_needs), 3)

    count = 0
    for things in thing_list:
        if things in ans_thing:
            count += 1
    ws.cell(row=2 + num, column=10).value = round(count / len(thing_list), 3)

    count = 0
    for ans_t in ans_thing:
        if ans_t in thing_list:
            count += 1
    ws.cell(row=2 + num, column=11).value = round(count / len(ans_thing), 3)

    ws.cell(row=2 + num, column=12).value = token_all

    # calculate room from things
    room_navi = []
    for t in thing_list:
        if t in situation['bathroom']:
            room_navi.append("bathroom")
        if t in situation['bedroom']:
            room_navi.append("bedroom")
        if t in situation['kitchen']:
            room_navi.append("kitchen")
        if t in situation['living room']:
            room_navi.append("living room")

    room_navi = set(room_navi)
    room_name = ",".join(room_navi)
    ws.cell(row=2 + num, column=6).value = room_name

    wb.save(EXCEL_NAME)


def chat_response(problem, needs_dict, wb, sheet, a, b):
    # LLM
    LLM = "gpt-4o-mini" if (b % 2 == 1) else "gpt-4o"

    # content
    if b % 3 == 1:
        content = content_murray
    elif b % 3 == 2:
        content = content_action
    else:
        content = content_explain

    # select needs
    content2 = (
        "Please select the needs that the user wants to satisfy in the above request.\n"
        "Please answer in the notation as listed, with no spaces, and only nouns.\n"
        "For nouns consisting of two words, insert a space.\n"
        + problem
    )
    chat = [
        {"role": "system", "content": content},
        {"role": "user", "content": content2},
    ]
    resp1 = client.chat.completions.create(model=LLM, messages=chat)
    msg1 = resp1.choices[0].message.content.lstrip()
    needs = msg1.split(",")
    print(msg1)

    # suggest destination based on needs
    needs_thing = []
    for k in needs_dict.keys():
        if k in needs:
            needs_thing += needs_dict[k]

    # dedup (optional but useful)
    needs_thing = list(dict.fromkeys([t.strip() for t in needs_thing if t.strip()]))

    print("thing list" + str(needs_thing))

    content3 = problem + """
\nBelow is a list of possible objects for the above request. Please narrow down and select only those that satisfy the user's request.
Please output your response in the following format.
Please do not include spaces, use only nouns, and answer in the notation provided.
For nouns consisting of two words, insert a space.
Also do not add [] or '.
Example is in the below.
Example:bed,sofa\n
""" + str(needs_thing)

    chat2 = [{"role": "user", "content": content3}]
    print("token count(user)：" + str(calc_token(content3)))

    resp2 = client.chat.completions.create(model=LLM, messages=chat2)
    msg2 = resp2.choices[0].message.content.lstrip()
    print(msg2)

    token_all = (
        calc_token(content)
        + calc_token(content2)
        + calc_token(content3)
        + calc_token(msg1)
        + calc_token(msg2)
    )
    print(token_all)

    write_xl(wb, msg1, msg2, token_all, sheet, a)


def main():
    a = 0
    b = 0
    Need_dict = {}

    sheets = [
        "test_murray_json", "test_murray_json_4o",
        "test_keyword_json", "test_keyword_json_4o",
        "test_explain_json", "test_explain_json_4o"
    ]

    wb = openpyxl.load_workbook(EXCEL_NAME)  # read Excel file

    for sheet in sheets:
        prob_list = [] 
        read_xl(wb, prob_list, Need_dict, sheet)
        b += 1
        a = 0
        for problem in prob_list:
            print(problem)
            chat_response(problem, Need_dict, wb, sheet, a, b)
            a += 1


if __name__ == "__main__":
    main()
